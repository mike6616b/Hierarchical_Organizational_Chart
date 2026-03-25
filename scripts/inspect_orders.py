import openpyxl
import json

file_path = '訂貨資料_DB.xlsx'

print("Loading workbook...")
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
print("Sheet names:", wb.sheetnames)

sheet = wb['rawdata']

headers = []
# header is on row 3 (index 2)
for idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if idx == 2:
        headers = list(row)
    if idx == 3:
        first_row = list(row)
        break

print("\n--- Headers ---")
for i, h in enumerate(headers):
    print(f"Col {i} ({openpyxl.utils.get_column_letter(i+1)}): {h}")

print("\n--- First Row ---")
for i, val in enumerate(first_row):
    print(f"{headers[i]}: {val}")
